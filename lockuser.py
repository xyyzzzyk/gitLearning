#!/usr/local/python27/bin/python3.7

import xlrd
import paramiko
import logging
import requests
import openpyxl
from openpyxl import load_workbook

# init log module
logger = logging.getLogger('connect')
logger.setLevel(logging.DEBUG)

fh = logging.FileHandler('exec.log')
fh.setLevel(logging.DEBUG)

ch = logging.StreamHandler()
ch.setLevel(logging.DEBUG)

formatter = logging.Formatter(
    '%(asctime)s - %(name)s - %(levelname)s - %(message)s')
fh.setFormatter(formatter)
ch.setFormatter(formatter)

logger.addHandler(fh)
logger.addHandler(ch)

# open excel


def open_excel(files):
    try:
        data = openpyxl.load_workbook(files)
        return data
    except Exception as e:
        print(str(e))


def ssh_connect(host, pwd, puser):      #connect and lock user
    flag = 1
    try:
        s = paramiko.SSHClient()
        paramiko.util.log_to_file('paramiko.log')
        s.set_missing_host_key_policy(paramiko.AutoAddPolicy())

        s.connect(host, port=22, username='ampmon',
              password=pwd, timeout=5, look_for_keys=False)
    
        r_command = "sudo passwd -l "+ puser
        r_testcmd = "echo 'Success " + r_command + "'"

        stdin, stdout, stderr = s.exec_command(r_command, get_pty=True)
    
   
        logger.info("[change_result] host: " + host.center(16) +
                            " ampmon/" + pwd.center(16) + r_command+ "........output: " + stdout.read().decode())
    except Exception as e:
        logger.error(host + "  with passwd: " + pwd.center(20) +"len:   "+str(len(pwd)).center(10) + str(e))
        flag = 0
    except paramiko.ssh_exception.AuthenticationException:
        logger.error("host:" + host + "  with passwd: " + pwd + " not match")
        flag = 0
    except paramiko.ssh_exception.SSHException:
        logger.error("host:" + host + " SSH connect fail   /"+str(e))
        flag = 0

    s.close()
    return flag

# 根据pip,查找密码表pfile EXCEL


def findpwd(pIP, pFile):

    pdata = open_excel(pFile)  # 打开密码表
    ptable = pdata.worksheets[0]
    prows = ptable.max_row
    pcols = ptable.max_column
    flag = 'none'
    for pvalue in range(prows):
        if pIP == ptable.cell(row=pvalue+1, column=7).value:
            flag = ptable.cell(row=pvalue+1, column=9).value

            break
    return flag

# 打开锁定用户excel，读取IP、用户，根据ip查找口令，连接主机，锁定用户


def lockUser(excelFile, passwdFile, b_col, c_col):
    data = open_excel(excelFile)
    table = data.worksheets[0]
    nrows = table.max_row
    ncols = table.max_column
    print(str(nrows) + "/" + str(ncols))
    for rdata in range(nrows):
        vhost = str(table.cell(row=rdata+1, column=b_col).value)
        vuser = str(table.cell(row=rdata+1, column=c_col).value)
        if len(vhost) > 6:
            vpwd = findpwd(vhost, passwdFile)


            if vpwd == 'none' :
                logger.error(vhost + "  find't passwd !!! Please check ip !")
                continue

            cflag = ssh_connect(vhost, vpwd, vuser)


def main(xFile, pwdFile):
    host_col = 6
    begin_row = 0
    pwd_col = 0
    lockUser(xFile, pwdFile, 1, 2)


if __name__ == "__main__":
    uFile = "lockuser.xlsx"
    pFile = "xyy.xlsx"

    main(uFile, pFile)
